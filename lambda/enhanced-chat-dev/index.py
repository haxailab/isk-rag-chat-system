import json
import boto3
import logging
import traceback
import os
import base64
from typing import Dict, Any, List, Optional
from datetime import datetime

# 構造化ログ設定
logger = logging.getLogger()
logger.setLevel(logging.DEBUG if os.getenv('LOG_LEVEL') == 'DEBUG' else logging.INFO)

# AWS クライアント初期化
s3_client = boto3.client('s3')
lambda_client = boto3.client('lambda')
bedrock_runtime = boto3.client('bedrock-runtime', region_name='ap-northeast-1')
bedrock_agent = boto3.client('bedrock-agent-runtime', region_name='ap-northeast-1')
dynamodb = boto3.resource('dynamodb')

# 設定
KNOWLEDGE_BASE_ID = os.getenv('KNOWLEDGE_BASE_ID', 'LK9Z59ROMF')
TEMP_FILES_BUCKET = os.getenv('TEMP_FILES_BUCKET')
MODEL_ID = os.getenv('CLAUDE_MODEL_ID', 'global.anthropic.claude-sonnet-4-6')
DOCUMENT_GENERATOR_FUNCTION = os.getenv('DOCUMENT_GENERATOR_FUNCTION', '')
ACCESS_LOG_TABLE = os.getenv('ACCESS_LOG_TABLE', '')
RAG_DOCUMENTS_BUCKET = 'isk-rag-documents-144828520862-ap-northeast-1'


def log_access(username: str, endpoint: str, session_id: str = None):
    """アクセスログをDynamoDBに記録（内容は保存しない・JST表示）"""
    if not ACCESS_LOG_TABLE:
        return
    try:
        from datetime import timezone, timedelta
        JST = timezone(timedelta(hours=9))
        table = dynamodb.Table(ACCESS_LOG_TABLE)
        now_utc = datetime.now(timezone.utc)
        now_jst = now_utc.astimezone(JST)
        table.put_item(Item={
            'username': username,
            'timestamp': now_jst.strftime('%Y-%m-%d %H:%M:%S JST'),
            'endpoint': endpoint,
            'session_id': session_id or '',
            'date': now_jst.strftime('%Y-%m-%d'),
            'ttl': int(now_utc.timestamp()) + 90 * 86400
        })
    except Exception as e:
        logger.warning(f"Access log write failed: {e}")


def wants_excel_generation(query: str) -> bool:
    """ユーザーがExcel/スプレッドシート生成を求めているか判定"""
    keywords = [
        'エクセル', 'excel', 'xlsx', 'xls', 'スプレッドシート',
        'spreadsheet', '表にして', '表に変換', '表形式',
    ]
    query_lower = query.lower().strip()
    return any(kw in query_lower for kw in keywords)


def generate_excel_from_session(session_id: str, query: str) -> Optional[Dict[str, Any]]:
    """セッション内のCSVデータからExcelファイルを生成"""
    if not session_id or not TEMP_FILES_BUCKET:
        return None

    try:
        # セッション内の抽出済みテキストからCSVデータを探す
        prefix = f"sessions/{session_id}/extracted/"
        response = s3_client.list_objects_v2(Bucket=TEMP_FILES_BUCKET, Prefix=prefix)

        csv_data = None
        csv_filename = None

        for obj in response.get('Contents', []):
            try:
                file_response = s3_client.get_object(Bucket=TEMP_FILES_BUCKET, Key=obj['Key'])
                content = file_response['Body'].read().decode('utf-8')
                extracted = json.loads(content)
                text = extracted.get('text', '')

                # CSVデータを含むファイルを探す
                if 'CSV Data' in text and 'rows' in text:
                    # "CSV Data (N rows, columns: ...):\n" の後のデータを取得
                    lines = text.split('\n')
                    # ヘッダー行（"CSV Data..."）をスキップしてCSV部分を取得
                    csv_lines = []
                    for line in lines:
                        if line.startswith('CSV Data'):
                            continue
                        if line.strip():
                            csv_lines.append(line)
                    if csv_lines:
                        csv_data = '\n'.join(csv_lines)

                        # オリジナルファイル名を取得
                        file_id = obj['Key'].split('/')[-1].replace('.json', '')
                        original_prefix = f"sessions/{session_id}/original/"
                        orig_response = s3_client.list_objects_v2(Bucket=TEMP_FILES_BUCKET, Prefix=original_prefix)
                        for orig_obj in orig_response.get('Contents', []):
                            if file_id in orig_obj['Key']:
                                head = s3_client.head_object(Bucket=TEMP_FILES_BUCKET, Key=orig_obj['Key'])
                                meta = head.get('Metadata', {})
                                if 'original-filename-b64' in meta:
                                    try:
                                        csv_filename = base64.b64decode(meta['original-filename-b64']).decode('utf-8')
                                    except Exception:
                                        csv_filename = meta.get('original-filename', 'data')
                                else:
                                    csv_filename = meta.get('original-filename', 'data')
                                break
                        break
            except Exception as e:
                logger.warning(f"Error reading session file: {e}")
                continue

        if not csv_data:
            return None

        # openpyxl で本物の .xlsx を生成
        import csv as csv_module
        import io
        import uuid
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        reader = csv_module.reader(csv_data.strip().split('\n'))
        rows = list(reader)

        if not rows:
            return None

        title = os.path.splitext(csv_filename or 'data')[0]

        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]

        # ISKブランドカラー
        header_fill = PatternFill(start_color='C8102E', end_color='C8102E', fill_type='solid')
        header_font = Font(bold=True, size=11, color='FFFFFF')
        default_font = Font(size=11)

        for row_idx, row in enumerate(rows, 1):
            for col_idx, cell_value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                else:
                    cell.font = default_font

        # 列幅を自動調整
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    # 日本語は幅2として計算
                    cell_len = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                    max_len = max(max_len, cell_len)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        # バッファに保存
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # S3にアップロード
        unique_id = str(uuid.uuid4())[:8]
        safe_title = "".join(c for c in title if c.isalnum() or c in "._- ").replace(" ", "_")[:50]
        filename = f"{safe_title}_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
        s3_key = f"documents/{unique_id}_{filename}"

        s3_client.put_object(
            Bucket=TEMP_FILES_BUCKET,
            Key=s3_key,
            Body=buffer.getvalue(),
            ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            ContentDisposition=f'attachment; filename="{filename}"'
        )

        # Presigned URL生成
        download_url = s3_client.generate_presigned_url(
            'get_object',
            Params={'Bucket': TEMP_FILES_BUCKET, 'Key': s3_key},
            ExpiresIn=900
        )

        logger.info(f"Excel generated: {s3_key}")
        return {
            'download_url': download_url,
            'filename': filename,
            'rows': len(rows),
            'title': title
        }

    except Exception as e:
        logger.error(f"Excel generation failed: {e}")
        return None

def search_internet(query: str, max_results: int = 3) -> List[Dict[str, Any]]:
    """インターネット検索機能（一時的に無効化）"""
    logger.info(f"Internet search temporarily disabled for query: {query}")
    # 一時的に空の結果を返す（requestsライブラリの問題を回避）
    return []


def enhanced_error_handler(func):
    """拡張エラーハンドリングデコレーター"""
    def wrapper(event, context):
        try:
            logger.info(f"Enhanced Chat Function start: {func.__name__}", extra={
                "requestId": context.aws_request_id,
                "functionName": context.function_name,
                "eventType": event.get('httpMethod', 'UNKNOWN')
            })

            result = func(event, context)

            logger.info("Enhanced Chat Function success", extra={
                "requestId": context.aws_request_id,
                "statusCode": result.get('statusCode')
            })
            return result

        except Exception as e:
            error_details = {
                "requestId": context.aws_request_id,
                "errorType": type(e).__name__,
                "errorMessage": str(e),
                "stackTrace": traceback.format_exc(),
                "inputEvent": json.dumps(event, default=str, ensure_ascii=False)
            }

            logger.error("Enhanced Chat Function error", extra=error_details)

            return {
                'statusCode': 500,
                'headers': get_cors_headers(),
                'body': json.dumps({
                    'error': 'Internal server error',
                    'requestId': context.aws_request_id,
                    'details': str(e) if os.getenv('LOG_LEVEL') == 'DEBUG' else 'チャット処理でエラーが発生しました',
                    'timestamp': datetime.utcnow().isoformat()
                }, ensure_ascii=False)
            }
    return wrapper


def get_cors_headers() -> Dict[str, str]:
    """CORS ヘッダーを取得"""
    return {
        'Content-Type': 'application/json',
        'Access-Control-Allow-Origin': '*',
        'Access-Control-Allow-Headers': 'Content-Type,Authorization,X-Amz-Date,X-Api-Key,X-Amz-Security-Token',
        'Access-Control-Allow-Methods': 'POST,OPTIONS',
        'Access-Control-Allow-Credentials': 'false'
    }


def search_knowledge_base(query: str, num_results: int = 10) -> Dict[str, Any]:
    """Knowledge Base検索（関連度スコアでフィルタリング）"""
    try:
        logger.info(f"Knowledge Base search started. Query: {query}, KB ID: {KNOWLEDGE_BASE_ID}")

        retrieve_response = bedrock_agent.retrieve(
            knowledgeBaseId=KNOWLEDGE_BASE_ID,
            retrievalQuery={'text': query}
        )

        retrieved_docs = []
        source_links = []
        MIN_RELEVANCE_SCORE = 0.5

        raw_results = retrieve_response.get('retrievalResults', [])
        logger.info(f"Knowledge Base retrieved {len(raw_results)} raw results")

        for result in raw_results:
            if 'content' in result and 'text' in result['content']:
                score = result.get('score', 0)
                if score < MIN_RELEVANCE_SCORE:
                    logger.info(f"Skipping low-score result: score={score:.3f}")
                    continue

                content = result['content']['text']
                retrieved_docs.append({
                    'content': content,
                    'source': 'knowledge_base',
                    'score': score
                })

                if 'location' in result and 's3Location' in result['location']:
                    source_uri = result['location']['s3Location']['uri']
                    if source_uri.startswith('s3://'):
                        filename = source_uri.split('/')[-1]
                        # ファイル名の拡張子を .pdf に変換（社内ファイルサーバー上の PDF を参照）
                        base_name, _ = os.path.splitext(filename)
                        pdf_filename = f"{base_name}.pdf"
                        # 社内ファイルサーバー UNC パス
                        file_link = f"\\\\Fileserv\\common\\部門間共有\\無機企画アクセスレベル2\\RAG_data\\{pdf_filename}"
                        source_links.append({
                            'filename': pdf_filename,
                            'url': file_link,
                            'score': score,
                            'source': 'knowledge_base'
                        })

        logger.info(f"Knowledge Base: {len(retrieved_docs)} relevant results (filtered from {len(raw_results)})")

        return {
            'success': True,
            'documents': retrieved_docs,
            'source_links': source_links,
            'total_results': len(retrieved_docs),
            'raw_results': len(raw_results)
        }

    except Exception as e:
        logger.error(f"Knowledge Base search failed: {e}")
        logger.error(f"Error type: {type(e).__name__}")

        # より具体的なエラー情報を提供
        error_details = str(e)
        if 'AccessDeniedException' in str(e):
            error_details = f"権限エラー: Knowledge Base '{KNOWLEDGE_BASE_ID}' へのアクセス権限がありません。IAM設定を確認してください。"
        elif 'ResourceNotFoundException' in str(e):
            error_details = f"Knowledge Base '{KNOWLEDGE_BASE_ID}' が見つかりません。IDが正しいか確認してください。"
        elif 'ValidationException' in str(e):
            error_details = f"リクエスト形式エラー: {str(e)}"

        return {
            'success': False,
            'documents': [],
            'source_links': [],
            'total_results': 0,
            'error': error_details
        }


def search_session_files(session_id: str, query: str) -> Dict[str, Any]:
    """セッション内ファイル検索（新規機能）"""
    if not session_id or not TEMP_FILES_BUCKET:
        return {
            'success': False,
            'documents': [],
            'source_links': [],
            'total_results': 0
        }

    try:
        logger.info(f"Session file search for session {session_id}, query: {query}")

        # セッション内の抽出済みテキストファイルを取得
        prefix = f"sessions/{session_id}/extracted/"

        response = s3_client.list_objects_v2(
            Bucket=TEMP_FILES_BUCKET,
            Prefix=prefix
        )

        session_documents = []
        session_source_links = []

        for obj in response.get('Contents', []):
            try:
                # 抽出済みテキストファイルを取得
                file_response = s3_client.get_object(
                    Bucket=TEMP_FILES_BUCKET,
                    Key=obj['Key']
                )

                content = file_response['Body'].read().decode('utf-8')
                extracted_data = json.loads(content)

                # より柔軟なファイル検索（セッション内の全ファイルを対象）
                extracted_text = extracted_data.get('text', '')

                # キーワード検索または画像/ファイル関連の質問の場合は全て含める
                should_include = False

                # 1. 直接的なキーワードマッチ
                if query.lower() in extracted_text.lower():
                    should_include = True

                # 2. ファイル関連の質問パターン
                file_related_patterns = [
                    # ファイル・アップロード関連
                    '画像', 'ファイル', 'アップロード', '見れる', '内容', 'jpg', 'png', 'pdf', 'txt',
                    # 文書・メモ関連
                    'メモ', '文書', 'ドキュメント', '資料', 'テキスト', '文章', '記録', 'ノート',
                    # 分析・処理関連
                    '要約', 'まとめ', '分析', '内容', '説明', '確認', '詳細', '情報',
                    # 質問詞
                    '何', 'どんな', 'どこ', 'いつ', 'なに', '写真', '図', '表', 'データ',
                    # その他
                    'この', 'それ', 'これ', '上記', '先ほど', 'さっき'
                ]
                for pattern in file_related_patterns:
                    if pattern in query.lower():
                        should_include = True
                        break

                # 3. OCR失敗の場合でも画像ファイルなら含める
                if 'OCR処理に失敗' in extracted_text and any(ext in extracted_text for ext in ['.jpg', '.png', '.jpeg', '.gif', '.webp']):
                    should_include = True

                # 4. セッション内ファイル数が少ない場合は積極的に含める（最大3ファイルまで）
                total_files = len(response.get('Contents', []))
                if total_files <= 3 and len(session_documents) < 3:
                    should_include = True

                if should_include:
                    # ファイルIDからオリジナルファイル名を取得
                    file_id = obj['Key'].split('/')[-1].replace('.json', '')

                    # オリジナルファイルの情報を取得
                    original_prefix = f"sessions/{session_id}/original/"
                    original_response = s3_client.list_objects_v2(
                        Bucket=TEMP_FILES_BUCKET,
                        Prefix=original_prefix
                    )

                    original_filename = None
                    for orig_obj in original_response.get('Contents', []):
                        if file_id in orig_obj['Key']:
                            # メタデータからファイル名を取得（Base64デコード対応）
                            head_response = s3_client.head_object(
                                Bucket=TEMP_FILES_BUCKET,
                                Key=orig_obj['Key']
                            )
                            metadata = head_response.get('Metadata', {})
                            if 'original-filename-b64' in metadata:
                                # Base64デコード
                                try:
                                    original_filename = base64.b64decode(metadata['original-filename-b64']).decode('utf-8')
                                except Exception:
                                    original_filename = metadata.get('original-filename', orig_obj['Key'].split('/')[-1])
                            else:
                                original_filename = metadata.get('original-filename', orig_obj['Key'].split('/')[-1])
                            break

                    session_documents.append({
                        'content': extracted_text[:3000],  # 3KB制限
                        'source': 'session_file',
                        'filename': original_filename or f'file-{file_id}',
                        'file_id': file_id,
                        'score': 1.0  # 簡単なスコアリング
                    })

                    # セッションファイルのダウンロードリンク（簡易版）
                    session_source_links.append({
                        'filename': original_filename or f'file-{file_id}',
                        'url': f"session-file://{session_id}/{file_id}",  # 実際のS3 URLは後で実装
                        'score': 1.0,
                        'source': 'session_file',
                        'file_id': file_id
                    })

            except Exception as file_error:
                logger.warning(f"Failed to process session file {obj['Key']}: {file_error}")

        logger.info(f"Session file search found {len(session_documents)} relevant documents")

        return {
            'success': True,
            'documents': session_documents,
            'source_links': session_source_links,
            'total_results': len(session_documents)
        }

    except Exception as e:
        logger.error(f"Session file search failed: {e}")
        return {
            'success': False,
            'documents': [],
            'source_links': [],
            'total_results': 0,
            'error': str(e)
        }


def build_hybrid_context(kb_results: Dict[str, Any], session_results: Dict[str, Any], user_message: str, internet_results: List[Dict[str, Any]] = None) -> Dict[str, Any]:
    """Knowledge Base、セッションファイル、インターネット検索結果を統合してプロンプトを構築"""

    # 全文書を統合
    all_documents = []
    all_source_links = []

    # Knowledge Base文書
    if kb_results.get('success'):
        all_documents.extend(kb_results['documents'])
        all_source_links.extend(kb_results['source_links'])

    # セッションファイル文書
    if session_results.get('success'):
        all_documents.extend(session_results['documents'])
        all_source_links.extend(session_results['source_links'])

    # インターネット検索結果
    internet_docs = []
    if internet_results:
        for i, result in enumerate(internet_results):
            if result.get('content'):
                internet_docs.append({
                    'content': result['content'],
                    'source': 'internet',
                    'title': result.get('title', f'インターネット検索結果 {i+1}'),
                    'url': result.get('url', ''),
                    'source_type': result.get('source', 'Web'),
                    'score': 0.5  # インターネット検索は中程度の優先度
                })
                all_source_links.append({
                    'title': result.get('title', f'インターネット検索結果 {i+1}'),
                    'source': f"🌐 インターネット検索 ({result.get('source', 'Web')})",
                    'url': result.get('url', ''),
                    'relevance_score': 0.5
                })
        all_documents.extend(internet_docs)

    # 文書の優先順位付け（スコア順）
    all_documents.sort(key=lambda x: x.get('score', 0), reverse=True)

    # 関連文書が多い場合は上位5件まで使用し、残りは件数を通知
    MAX_DOCS = 5
    top_documents = all_documents[:MAX_DOCS]
    remaining_docs = len(all_documents) - len(top_documents)
    # KB検索の生の結果数（スコアフィルタ前）
    total_raw_kb = kb_results.get('raw_results', 0) if kb_results.get('success') else 0

    # 共通のベースプロンプト
    BASE_SYSTEM_PROMPT = """あなたは「無機RAG」という名前のISK（石原産業株式会社）社内向けAIアシスタントです。
ISKは「Local Insight, Global Impact」をスローガンとする無機化学品（酸化チタン等）メーカーです。

## 回答の基本ルール
1. 提供された<context>内の情報を最優先で使用する
2. contextに根拠がある場合は「資料によると〜」と明示する
3. contextに情報がない場合は「資料には記載がありません。一般的には〜」と明確に区別する
4. 推測で補う場合は必ず「推測ですが〜」と前置きする
5. 不確かな情報を断定しない

## 回答スタイル
- 丁寧で親しみやすい日本語
- 結論を先に、理由・詳細を後に
- 長い回答は箇条書きや見出しで整理
- 技術的質問には正確さを優先

## ISKについて
- 事業：酸化チタン・無機化学品の製造販売
- 拠点：三重県四日市市
- 担当者が社内文書を基に業務判断するための補助AIとして機能する

## Excel生成機能
ユーザーが「エクセルにして」等と依頼した場合、システムが自動生成します。「ファイル生成できません」とは言わず「Excelファイルを生成しました。下のリンクからダウンロードできます。」と回答してください。"""

    if not top_documents:
        # 文書が見つからない場合は一般知識モード
        return {
            'has_documents': False,
            'system_prompt': BASE_SYSTEM_PROMPT + "\n\n## 注意\n社内文書の検索結果が0件でした。一般知識から回答しますが、回答の最後に「※この回答は一般知識に基づいており、ISK社内文書は参照していません」と明記してください。",
            'source_links': [],
            'is_hybrid_response': False,
            'model_description': "Claude Sonnet（一般知識）"
        }

    # 文書コンテキストを構築
    kb_docs = [doc for doc in top_documents if doc['source'] == 'knowledge_base']
    session_docs = [doc for doc in top_documents if doc['source'] == 'session_file']
    internet_docs_filtered = [doc for doc in top_documents if doc['source'] == 'internet']

    context_parts = []

    if kb_docs:
        # 同じファイルのチャンクを結合
        merged_kb = {}
        for doc in kb_docs:
            # ソースリンクからファイル名を取得
            fn = 'unknown'
            for link in all_source_links:
                if link.get('source') == 'knowledge_base' and link.get('score') == doc.get('score'):
                    fn = link.get('filename', 'unknown')
                    break
            if fn not in merged_kb:
                merged_kb[fn] = doc['content'][:1500]
            else:
                merged_kb[fn] += '\n' + doc['content'][:1000]
        kb_context = "\n\n".join([f"【文書: {fn}】\n{content[:3000]}" for fn, content in merged_kb.items()])
        context_parts.append(f"=== ISK社内Knowledge Base ===\n{kb_context}")

    if session_docs:
        session_context = "\n\n".join([f"【セッションファイル{i+1}: {doc['filename']}】\n{doc['content']}" for i, doc in enumerate(session_docs)])
        context_parts.append(f"=== セッション内アップロードファイル ===\n{session_context}")

    if internet_docs_filtered:
        internet_context = "\n\n".join([f"【インターネット検索結果{i+1}】\nタイトル: {doc['title']}\nソース: {doc.get('source_type', 'Web')}\n内容: {doc['content']}" for i, doc in enumerate(internet_docs_filtered)])
        context_parts.append(f"=== 🌐 インターネット検索情報 ===\n{internet_context}")

    context_text = "\n\n".join(context_parts)

    # ユニークなKBファイル数を計算
    kb_filenames = set()
    for link in all_source_links:
        if link.get('source') == 'knowledge_base' and link.get('filename'):
            kb_filenames.add(link['filename'])
    unique_kb_count = len(kb_filenames) if kb_filenames else len(kb_docs)
    logger.info(f"Unique KB files: {unique_kb_count}, filenames: {kb_filenames}, kb_docs: {len(kb_docs)}")

    # ハイブリッドプロンプトを構築
    source_summary_parts = []
    if unique_kb_count > 0:
        source_summary_parts.append(f"社内KB {unique_kb_count}件")
    if len(session_docs) > 0:
        source_summary_parts.append(f"アップロードファイル {len(session_docs)}件")
    if len(internet_docs_filtered) > 0:
        source_summary_parts.append(f"インターネット {len(internet_docs_filtered)}件")
    source_summary = '、'.join(source_summary_parts)

    system_prompt = BASE_SYSTEM_PROMPT + f"""

## 今回の検索結果
参照可能な資料: {source_summary}
{f'※さらに{remaining_docs}件の関連資料がありますが、上位{len(top_documents)}件を参照しています。' if remaining_docs > 0 else ''}

## 情報源の明示ルール
- 社内KB由来 →「資料によると〜」
- アップロードファイル由来 →「アップロードされたファイルによると〜」
- インターネット由来 →「🌐インターネット情報：〜」と明確に区別
- contextに該当情報がない →「資料には記載がありません。一般的には〜」

回答の最後に「※参照: {source_summary}」と1行で明記してください。"""

    return {
        'has_documents': True,
        'system_prompt': system_prompt,
        'context_text': context_text,
        'source_links': all_source_links,
        'is_hybrid_response': True,
        'model_description': f"Hybrid RAG対応 Claude Sonnet (KB:{unique_kb_count} + Session:{len(session_docs)} + Internet:{len(internet_docs_filtered)})",
        'knowledge_base_count': unique_kb_count,
        'session_file_count': len(session_docs),
        'internet_search_count': len(internet_docs_filtered)
    }


@enhanced_error_handler
def handler(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    """Enhanced Chat ハンドラー（5分タイムアウト対応・タイプライター表示）"""

    # CORS Preflight
    if event.get('httpMethod') == 'OPTIONS':
        return {'statusCode': 200, 'headers': get_cors_headers(), 'body': '{}'}

    # 認証情報取得（Cognito authorizer）
    auth_info = event.get('requestContext', {}).get('authorizer', {})
    user_name = auth_info.get('claims', {}).get('cognito:username', 'anonymous')

    # リクエストボディ解析
    try:
        body = json.loads(event.get('body', '{}'))
        user_message = body.get('message', '')
        session_id = body.get('session_id')

        if not user_message:
            return {
                'statusCode': 400,
                'headers': get_cors_headers(),
                'body': json.dumps({'error': 'Message is required'}, ensure_ascii=False)
            }

        logger.info(f"Enhanced Chat from user: {user_name}, message: {user_message[:100]}...")
        log_access(user_name, 'enhanced-chat', session_id)

    except Exception as e:
        return {
            'statusCode': 400,
            'headers': get_cors_headers(),
            'body': json.dumps({'error': 'Invalid request', 'details': str(e)}, ensure_ascii=False)
        }

    # Knowledge Base + セッションファイル検索
    try:
        kb_results = search_knowledge_base(user_message)
        internet_results = search_internet(user_message, max_results=3)
        session_results = {'success': False, 'documents': [], 'source_links': [], 'total_results': 0}
        if session_id:
            session_results = search_session_files(session_id, user_message)

        hybrid_context = build_hybrid_context(kb_results, session_results, user_message, internet_results)

    except Exception as e:
        logger.error(f"Search failed: {e}\n{traceback.format_exc()}")
        hybrid_context = {
            'has_documents': False,
            'system_prompt': f"あなたは「無機RAG」という名前のISK社内向けAIアシスタントです。\nKB検索エラー: {str(e)}\n一般知識から回答してください。",
            'source_links': [],
            'is_hybrid_response': False,
            'model_description': "Claude Sonnet（一般知識）"
        }

    # Bedrock ストリーミング呼び出し（チャンクを集約）
    try:
        # <context>/<question> 構造化テンプレートでユーザーメッセージを構築
        context_text = hybrid_context.get('context_text', '')
        if context_text:
            structured_message = f"""<context>
{context_text}
</context>

<question>
{user_message}
</question>

上記のcontextを参考に、questionに答えてください。"""
        else:
            structured_message = user_message

        response = bedrock_runtime.invoke_model_with_response_stream(
            modelId=MODEL_ID,
            body=json.dumps({
                'anthropic_version': 'bedrock-2023-05-31',
                'max_tokens': 4096,
                'system': hybrid_context['system_prompt'],
                'messages': [{'role': 'user', 'content': structured_message}],
                'temperature': 0.7
            })
        )

        answer = ''
        stream_body = response.get('body')
        for stream_event in stream_body:
            chunk = stream_event.get('chunk')
            if chunk:
                chunk_data = json.loads(chunk.get('bytes').decode('utf-8'))
                if chunk_data.get('type') == 'content_block_delta':
                    text = chunk_data.get('delta', {}).get('text', '')
                    if text:
                        answer += text
                elif chunk_data.get('type') == 'message_stop':
                    break

    except Exception as e:
        logger.error(f"Bedrock call failed: {e}")
        return {
            'statusCode': 500,
            'headers': get_cors_headers(),
            'body': json.dumps({'error': f'AI呼び出しエラー: {str(e)}'}, ensure_ascii=False)
        }

    # Excel生成チェック
    document_url = None
    document_filename = None
    if wants_excel_generation(user_message) and session_id:
        excel_result = generate_excel_from_session(session_id, user_message)
        if excel_result:
            document_url = excel_result['download_url']
            document_filename = excel_result['filename']

    # ソースリンク重複排除
    is_rag = hybrid_context.get('has_documents', False)
    raw_links = hybrid_context.get('source_links', []) if is_rag else []
    seen = set()
    effective_source_links = []
    for link in raw_links:
        fn = link.get('filename', '')
        if fn and fn not in seen:
            seen.add(fn)
            effective_source_links.append(link)

    return {
        'statusCode': 200,
        'headers': get_cors_headers(),
        'body': json.dumps({
            'answer': answer,
            'source_links': effective_source_links,
            'model': f"無機RAG powered by {hybrid_context.get('model_description', 'Claude Sonnet')}",
            'is_rag_response': is_rag,
            'is_hybrid_response': hybrid_context.get('is_hybrid_response', False),
            'session_id': session_id,
            'document_url': document_url,
            'document_filename': document_filename,
            'search_stats': {
                'knowledge_base_results': kb_results.get('total_results', 0) if 'kb_results' in locals() else 0,
                'session_file_results': session_results.get('total_results', 0) if 'session_results' in locals() else 0,
                'total_sources': len(hybrid_context.get('source_links', []))
            }
        }, ensure_ascii=False)
    }